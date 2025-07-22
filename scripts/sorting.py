# Импортируем необходимые библиотеки для работы с Excel, конфигами, HTTP-запросами и файловой системой

import pandas as pd                                # Работа с таблицами Excel
import toml                                        # Чтение конфигурационных файлов TOML
import os                                          # Операции с файловой системой
import requests                                    # HTTP-запросы для загрузки описаний и изображений
from datetime import datetime                      # Работа с датой и временем
from openpyxl import load_workbook                 # Работа с Excel-файлами (автоподгонка ширины столбцов)
from openpyxl.utils import get_column_letter       # Получение буквенного обозначения столбца Excel
from concurrent.futures import ThreadPoolExecutor  # Параллельная загрузка данных (многопоточность)
import threading                                   # Потокобезопасное логирование
import csv                                         # Работа с CSV-файлами (логирование отсутствующих изображений)

# Уровень модуля для потокобезопасного логирования
_log_file = "missing_images.csv"  # Файл для логирования отсутствующих изображений
_log_lock = threading.Lock()      # Блокировка для потокобезопасной записи


# Класс для сортировки и обработки Excel-файлов
class ExcelSorter: 
    def __init__(self, file_path_save, progress_callback=None):
        # Формируем путь для сохранения файла с учетом текущей даты
        self.progress_callback = progress_callback
        current_date = datetime.now().strftime("%d-%m-%Y")

        if file_path_save == "":
            self.file_path_save = os.path.join(os.path.expanduser("~"), "Desktop", f"filtered_{current_date}.xlsx")
        else:
            if not file_path_save.endswith('.xlsx'):
                file_path_save += f"/filtered_{current_date}.xlsx"
            self.file_path_save = file_path_save
            
        self.config_path = "config.toml"
    
    def sortExcel(self, lieferung_value, aufschlag_value1, aufschlag_value2, file_path, checkbox_kaufpreis, checkbox_steuer):
        # Загрузка конфигурации из TOML-файла
        config = toml.load(self.config_path)

        # Используем значения по умолчанию, если не переданы параметры
        if lieferung_value == "":
            lieferung_value = config['default']['lieferung_value']
        if aufschlag_value1 == "":
            aufschlag_value1 = config['default']['aufschlag_value1']
        if aufschlag_value2 == "":
            aufschlag_value2 = config['default']['aufschlag_value2']

        # Преобразуем значения в числовой тип
        lieferung_value = float(lieferung_value)
        aufschlag_value1 = float(aufschlag_value1)
        aufschlag_value2 = float(aufschlag_value2)
        
        # Чтение значения первой ячейки и проверка
        df = pd.read_excel(file_path)

        # Безопасное извлечение значения первой ячейки
        first_cell_value = df.iat[0, 0]  # Используем .iat для доступа к первой ячейке
        print(f"Значение первой ячейки: '{first_cell_value}'")
        
        # Преобразуем значение в строку для корректного сравнения
        first_cell_value = str(first_cell_value).strip() if first_cell_value is not None else None

        # В зависимости от значения первой ячейки вызываем нужный метод сортировки
        if  first_cell_value == "Brands":
            print("Условие выполнено: первая ячейка пуста")
            self.methodSortZwei(aufschlag_value1, file_path, checkbox_kaufpreis, checkbox_steuer)
        else:
            print(f"Условие не выполнено: первая ячейка содержит '{first_cell_value}'")
            self.methodSortEins(lieferung_value, aufschlag_value1, aufschlag_value2, file_path, checkbox_kaufpreis, checkbox_steuer)
    
    def calculate(  
                self,                       
                preis,
                lieferung_value, 
                aufschlag_value1,
                aufschlag_value2, 
                checkbox_steuer
                ):
        # Расчет итоговой цены с учетом надбавок и налога
        STEUER = 1.19
        if preis < 500:
            preis = preis + aufschlag_value1 + lieferung_value
        else:
            preis = preis + (preis * aufschlag_value2 / 100) + lieferung_value

        if checkbox_steuer:
            preis *= STEUER

        return round(preis, 2)

    def update_progress_bar(self, current_step, total_steps):
        if self.progress_callback:
            self.progress_callback(current_step, total_steps)

    def methodSortEins(self, lieferung_value, aufschlag_value1, aufschlag_value2, file_path, checkbox_kaufpreis, checkbox_steuer, export_to_csv=True):
        # Основная сортировка для обычных файлов
        df = pd.read_excel(file_path)
        config = toml.load(self.config_path)
        bezeichnung_filter = config['dataSet']['data']

        # Оставляем только нужные столбцы и фильтруем по первым словам
        df = df[["Artikelnummer", "Bezeichnung", "Preis", "Zustand"]].copy()
        df['Artikelnummer_raw'] = df['Artikelnummer'].astype(str)
        df['Artikelnummer'] = df['Artikelnummer'].astype(str) + df['Zustand'].astype(str).str.upper()
        df['First_Word'] = df['Bezeichnung'].str.split().str[0]
        df = df[df['First_Word'].isin(bezeichnung_filter)].drop(columns=['First_Word'])
        
        # Функция для получения URL изображения
        def get_bilder_url(artikelnummer_raw, bezeichnung):
            base_url = "https://telefoneria.com/wp-content/uploads/products/bilder"
            default_url = f"{base_url}/standart.webp"

            # 1. Попытка артикульного изображения
            article_url = f"{base_url}/{artikelnummer_raw}.webp"
            try:
                resp = requests.head(article_url, timeout=100)
                if resp.status_code == 200:
                    print(f"\033[92mBild vorhanden für {artikelnummer_raw}: {article_url}\033[0m")
                    return article_url
                else:
                    print(f"\033[91mKein Bild für {artikelnummer_raw} (Status {resp.status_code})\033[0m")
                    # Логируем отсутствие
                    with _log_lock, open(_log_file, "a", newline="", buffering=1) as csvfile:
                        writer = csv.writer(csvfile)
                        writer.writerow([artikelnummer_raw, bezeichnung])
            except Exception as e:
                print(f"\033[91mFehler beim Zugriff auf {article_url}: {e}\033[0m")
                with _log_lock, open(_log_file, "a", newline="", buffering=1) as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow([artikelnummer_raw, bezeichnung])

            # 2. Фоллбек по производителю
            name_lower = bezeichnung.lower()
            for manufacturer in ("apple", "samsung", "xiaomi"):
                if manufacturer in name_lower:
                    manu_url = f"{base_url}/standart_{manufacturer}.webp"
                    try:
                        resp = requests.head(manu_url, timeout=100)
                        if resp.status_code == 200:
                            print(f"\033[93mHerstellerbild gefunden ({manufacturer}): {manu_url}\033[0m")
                            return manu_url
                        else:
                            print(f"\033[91mKein Herstellerbild für {manufacturer} (Status {resp.status_code})\033[0m")
                    except Exception as e:
                        print(f"\033[91mFehler beim Zugriff auf {manu_url}: {e}\033[0m")
                    break

            # 3. Генеральный фоллбек
            print(f"\033[91mVerwende generisches Standardbild: {default_url}\033[0m")
            return default_url

        # Функция для загрузки полной описания
        def beschreibung_abrufen(artikelnummer_raw):
            url = f"https://telefoneria.com/wp-content/uploads/products/beschreibung/{artikelnummer_raw}.html"
            try:
                resp = requests.get(url, timeout=100) 
                resp.encoding = 'utf-8'
                if resp.status_code == 200:
                    print(f"\033[92mBeschreibung geladen für {artikelnummer_raw}\033[0m")
                    return resp.text.replace('\n', '').replace('\r', '')
                else:
                    print(f"\033[91mFehler beim Laden der Datei {url}: {resp.status_code}\033[0m")
            except Exception as e:
                print(f"\033[91mFehler beim Zugriff auf {url}: {e}\033[0m")
            return " "  # При ошибке или отсутствии файла возвращаем пробел

        # Функция для загрузки короткого описания
        def kurze_beschreibung_abrufen(artikelnummer_raw):
            url = f"https://telefoneria.com/wp-content/uploads/products/kurzebeschreibung/{artikelnummer_raw}.html"
            try:
                resp = requests.get(url, timeout=100)  
                resp.encoding = 'utf-8'
                if resp.status_code == 200:
                    print(f"\033[92mBeschreibung geladen für {artikelnummer_raw}\033[0m")
                    return resp.text.replace('\n', '').replace('\r', '')
                else:
                    print(f"\033[91mFehler beim Laden der Datei {url}: {resp.status_code}\033[0m")
            except Exception as e:
                print(f"\033[91mFehler beim Zugriff auf {url}: {e}\033[0m")
            return " "  # При ошибке или отсутствии файла возвращаем пробел
        
        # Параллельная загрузка данных (без прогресса)
        def fetch_all_beschreibungen_parallel(funktion, artikelnummer_list, max_threads=60):
            with ThreadPoolExecutor(max_workers=max_threads) as executor:
                return list(executor.map(funktion, artikelnummer_list))

        # Параллельная загрузка данных с прогресс-баром
        def fetch_all_beschreibungen_parallel_with_progress(funktion, artikelnummer_list, total_steps, start_step=0, max_threads=4):
            results = [None] * len(artikelnummer_list)
            futures = []
            with ThreadPoolExecutor(max_workers=max_threads) as executor:
                for idx, item in enumerate(artikelnummer_list):
                    futures.append((idx, executor.submit(funktion, item)))
                finished = 0
                for idx, future in futures:
                    result = future.result()  # Ждём результат
                    results[idx] = result
                    finished += 1
                    self.update_progress_bar(start_step + finished, total_steps)
            return results

        # Добавляем столбцы с изображениями и описаниями с отслеживанием прогресса
        artikelnummer_list = df['Artikelnummer_raw'].tolist()
        total_steps = len(artikelnummer_list) * 3
        step = 0

        df['Bilder'] = fetch_all_beschreibungen_parallel_with_progress(
            lambda artikelnummer_raw: get_bilder_url(
                artikelnummer_raw, 
                df.loc[df['Artikelnummer_raw'] == artikelnummer_raw, 'Bezeichnung'].values[0]
            ),
            artikelnummer_list, total_steps, step
        )
        step += len(artikelnummer_list)

        df['Beschreibung'] = fetch_all_beschreibungen_parallel_with_progress(
            beschreibung_abrufen, artikelnummer_list, total_steps, step
        )
        step += len(artikelnummer_list)

        df['Kurzbeschreibung'] = fetch_all_beschreibungen_parallel_with_progress(
            kurze_beschreibung_abrufen, artikelnummer_list, total_steps, step
        )

        # Преобразуем цены в числовой формат
        df['Preis'] = pd.to_numeric(df['Preis'].astype(str).str.replace('€', '').str.replace(' ', ''), errors='coerce')

        # Добавляем столбец "Kaufpreis" при необходимости
        if checkbox_kaufpreis:
            df['Kaufpreis'] = df['Preis']

        # Применяем расчет цены
        def apply_calc(price):
            if pd.notna(price):
                return self.calculate(price, lieferung_value, aufschlag_value1, aufschlag_value2, checkbox_steuer)
            return price

        df['Preis'] = df['Preis'].apply(apply_calc)
        df['Preis'] = df['Preis'].apply(lambda x: f"{x:.2f} €" if pd.notna(x) else x)

        # Определяем категорию по названию
        def get_kategorie_from_bezeichnung(bezeichnung: str) -> str:
            name = bezeichnung.lower()

            # Определяем категорию для Apple по ключевым словам
            if "apple" in name:
                if "ipad" in name:
                    return "Tablet > Apple"         # Если в названии есть "ipad"
                elif "iphone" in name:
                    return "Handy > Apple"          # Если в названии есть "iphone"
                elif "macbook" in name:
                    return "Notebook > Apple"       # Если в названии есть "macbook"
                elif "watch" in name:
                    return "Smartwatch > Apple"     # Если в названии есть "watch"
                elif "airpods" in name:
                    return "Kopfhörer > Apple"      # Если в названии есть "airpods"

            # Категории для Samsung
            elif "samsung" in name:
                if "tab" in name:
                    return "Tablet > Samsung"       # Если в названии есть "tab"
                elif "galaxy" in name and "watch" not in name and "tab" not in name and "buds" not in name:
                    return "Handy > Samsung"        # Если "galaxy", но не "watch", "tab", "buds"
                elif any(x in name for x in ["chromebook", "notebook", "book"]):
                    return "Notebook > Samsung"     # Если есть "chromebook", "notebook", "book"
                elif "watch" in name:
                    return "Smartwatch > Samsung"   # Если есть "watch"
                elif "buds" in name:
                    return "Kopfhörer > Samsung"    # Если есть "buds"

            # Категории для Xiaomi
            elif "xiaomi" in name:
                if "tab" in name:
                    return "Tablet > Xiaomi"        # Если в названии есть "tab"
                elif not any(x in name for x in ["tab", "chromebook", "notebook", "book", "watch", "buds"]):
                    return "Handy > Xiaomi"         # Если нет других ключевых слов
                elif any(x in name for x in ["chromebook", "notebook", "book"]):
                    return "Notebook > Xiaomi"      # Если есть "chromebook", "notebook", "book"
                elif "watch" in name:
                    return "Smartwatch > Xiaomi"    # Если есть "watch"
                elif "buds" in name:
                    return "Kopfhörer > Xiaomi"     # Если есть "buds"

            # Категория по умолчанию
            return "Sonstiges"                      # Если не найдено ни одно из условий

        df['Kategorien'] = df['Bezeichnung'].apply(get_kategorie_from_bezeichnung)
        # Не добавляем состояние к имени
        # df['Bezeichnung'] = df['Bezeichnung'].astype(str) + " Zustand:" + df['Zustand'].astype(str).str.capitalize()
        df = df.rename(columns={'Bezeichnung': 'Name'})

        # Переименовываем zustand в Schlagwörter
        df = df.rename(columns={'Zustand': 'Schlagwörter'})
        df['Schlagwörter'] = df['Schlagwörter'].astype(str).str.capitalize()

        # Удаляем временный столбец
        df.drop(columns=['Artikelnummer_raw'], inplace=True)

        # Сохраняем в Excel или CSV
        try:
            if export_to_csv:
                # Сохраняем в CSV с запятыми
                csv_path = self.file_path_save.replace('.xlsx', '.csv')
                df.to_csv(csv_path, index=False, sep=',')
                print(f"Файл CSV сохранён: {csv_path}")
            else:
                # Сохраняем в Excel
                with pd.ExcelWriter(self.file_path_save, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False)
                self.autofit_columns(self.file_path_save)
                print(f"Файл Excel сохранён: {self.file_path_save}")

        except Exception as e:
            print(f"Ошибка при сохранении: {e}")
            backup_path = os.path.join(os.path.expanduser("~"), "Desktop",
                                    f"filtered_example_backup.{ 'csv' if export_to_csv else 'xlsx' }")
            if export_to_csv:
                df.to_csv(backup_path, index=False, sep=',')
            else:
                df.to_excel(backup_path, index=False)
                self.autofit_columns(backup_path)
            print(f"Резервный файл сохранён: {backup_path}")

    def methodSortZwei(self, aufschlag_value1, file_path, checkbox_kaufpreis, checkbox_steuer):
        # Сортировка для файлов с особыми заголовками (например, если первая ячейка "Brands")
        df = pd.read_excel(file_path)
        
        # Создаем новый DataFrame с реальными заголовками из второй строки
        headers = df.iloc[0]
        filtered_df = df.iloc[1:].copy()
        filtered_df.columns = headers
        filtered_df.reset_index(drop=True, inplace=True)
        
        # Если checkbox_kaufpreis равно True, добавляем столбец 'Kaufpreis'
        if checkbox_kaufpreis:
            # Сохраняем оригинальные цены перед преобразованием
            filtered_df['Kaufpreis'] = filtered_df['Selling_Price'].copy()

        # Преобразуем цены в числовой формат
        filtered_df['Selling_Online_Price'] = pd.to_numeric(
            filtered_df['Selling_Online_Price'].astype(str).str.replace('€', '').str.replace(' ', ''),
            errors='coerce'
        )
        filtered_df['Selling_Price'] = pd.to_numeric(
            filtered_df['Selling_Price'].astype(str).str.replace('€', '').str.replace(' ', ''),
            errors='coerce'
        )
        
        # Создаем новый столбец для форматированных цен
        filtered_df['Formatted_Preis'] = filtered_df['Selling_Price'].copy()

        if 'Selling_Price' in filtered_df.columns:
            for index, row in filtered_df.iterrows():
                price = row['Selling_Price']
                
                if pd.notna(price):
                    formatted_price = self.calculate(
                        price,
                        0,
                        aufschlag_value1,
                        0,
                        checkbox_steuer
                    )
                    
                    filtered_df.at[index, 'Formatted_Preis'] = f"{formatted_price:.2f} €"

        # Заменяем столбец 'Preis' на форматированный
        filtered_df['Selling_Price'] = filtered_df['Formatted_Preis']
        filtered_df.drop(columns=['Formatted_Preis'], inplace=True)

        # Добавляем символ евро к столбцу Selling_Online_Price
        filtered_df['Selling_Online_Price'] = filtered_df['Selling_Online_Price'].astype(str) + ' €'
        filtered_df.reset_index(drop=True, inplace=True)

        try:
            # Сохраняем результат через ExcelWriter
            with pd.ExcelWriter(self.file_path_save, engine='openpyxl') as writer:
                filtered_df.to_excel(writer, index=False, header=True)
            
            # Подгоняем размеры столбцов
            self.autofit_columns(self.file_path_save)
            
            print(f"Фильтрация завершена. Новый файл сохранён по пути: {self.file_path_save}")
        except Exception as e:
            print(f"Ошибка при сохранении файла: {str(e)}")
            # Попробуем сохранить на рабочий стол, если возникла ошибка
            backup_path = os.path.join(os.path.expanduser("~"), "Desktop", "filtered_example_backup.xlsx")
            filtered_df.to_excel(backup_path, index=False, header=True)
            self.autofit_columns(backup_path)
            print(f"Файл сохранен в резервное расположение: {backup_path}")

    def autofit_columns(self, filename):
        """Автоматическая подгонка ширины столбцов"""
        workbook = load_workbook(filename)
        worksheet = workbook.active

        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        workbook.save(filename)

    def read_first_column_value(self, file_path):
        """
        Читает значение первой строки первого столбца Excel файла.
        Если первая строка пустая, возвращает значение второй строки.
        """
        try:
            # Читаем Excel файл
            df = pd.read_excel(file_path)
            
            # Получаем первый столбец
            first_column = df.iloc[:, 0]
            
            # Проверяем первую строку
            if pd.notna(first_column.iloc[0]):
                return first_column.iloc[0]
            
            # Если первая строка пустая, возвращаем вторую строку
            if len(first_column) > 1 and pd.notna(first_column.iloc[1]):
                return first_column.iloc[1]
        
        except Exception as e:
            print(f"Ошибка при чтении файла: {e}")
